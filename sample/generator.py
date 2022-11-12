#!/usr/bin/python3

# handling generator, which is made from a part of
# column or row that is come from sheet of xlsx.
import os
import platform

SYSTEM=platform.system()
MACHINE=platform.machine()
#작업폴더 설정
working_dir=os.getcwd()
if SYSTEM == "Windows":
    os.chdir("D:\개발_코딩연습\python_연습\데이터분석입문\ch3")
    pass
elif SYSTEM == "macOS":
    pass
elif SYSTEM== "Linux":
    if MACHINE == 'aarch64':    #smart phone
        os.chdir("/storage/emulated/0/Documents/py3_phone/sample/")
        

#getting a generator, containing datetime.
import openpyxl

xl_file = "sales_2013.xlsx"
wb=openpyxl.load_workbook( xl_file )
ws=wb.active

for cell in ws[ '2' ]:    #2행에서 날짜셀 찾기
    if cell.data_type == 'd':
        idx_colu_date = cell.column    #정수형
        print( idx_colu_date, cell.column_letter )
        
#a=ws['a3']
#print( dir(a) )    #finding cell methods

#making one generator
datetime_gen= ws.iter_cols( min_row=2,
                                                      min_col=idx_colu_date,
                                                      max_col=idx_colu_date
                                                    )
c=next( datetime_gen )
print("==="*10,"\n", c)

print("printing methods, generator type obj -datetime_gen-")
print( type(datetime_gen),  '\n' ,
#            len(datetime_gen),   '\n',   #not working.
              len( c ), '\n',
             dir(datetime_gen),    '\n' ,
             
          )
          
#U can check that generator is made by 
# one time && one object.
cnt=1
for i in datetime_gen:
    print( cnt, '\t', i   )
    cnt += 1

#tmp = next(datetime_gen)    
'''
Have you ever had to work with 
##a dataset so large that it overwhelmed 
your machine’s memory?
 ##Or maybe you have ##a complex function
 that needs to maintain an internal state 
every time it’s called, but the function i
s too small to justify creating its own class##.
 In these cases and more, generators and 
the Python yield statement are here to help.
ref) https://realpython.com/introduction-to-python-generators/
'''