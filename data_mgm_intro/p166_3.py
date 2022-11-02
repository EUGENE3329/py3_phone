#!/usr/bin/python3

import os
from re import T
import sys

from datetime import date

import openpyxl

os.system('clear')

#작업 디렉토리가 실행파일 장소와 일치하지 않아서 생기는 에러 수정
working_dir=os.getcwd()
print('work here: ', working_dir )

'''pc경로
if working_dir != "D:\개발_코딩연습\python_연습\데이터분석입문\ch3":
    os.chdir("D:\개발_코딩연습\python_연습\데이터분석입문\ch3")
print("Now !!!: ", os.getcwd())
'''

'''스마트폰 경로
/storage/emulated/0/Documents/py3/data_mgm_intro/p162_1.py
'''
if working_dir != "/storage/emulated/0/Documents/py3_phone/data_mgm_intro/":
    os.chdir("/storage/emulated/0/Documents/py3_phone/data_mgm_intro/")
print("Now !!!: ", os.getcwd())

#input_xl_f = sys.argv[1]
#input_xl_f = sys.argv[1]
input_xl_f =  'sales_2013.xlsx'
output_xl_f = 'tmp3.xlsx'

output_wb = openpyxl.Workbook()
#output_wb.create_sheet('jan_2013_output')
#output_ws1=output_wb.get_sheet_by_name('jan_2013_output')
output_ws1=output_wb.active
output_ws1.title='jan_2013_output'


#경고 삭제, 아래 순환문 때문에 작성 함.
#폰 에서 코딩 위치
import warnings
warnings.simplefilter('ignore')

#wb=openpyxl.load_workbook(input_xl_f, read_only=True, data_only=True)
wb=openpyxl.load_workbook(input_xl_f)
ws1=wb['january_2013']
''
'''
#pc에서 코딩 위치
import warnings
warnings.simplefilter('ignore')
'''

# openpyxl공식 문서 추천, 값만 추출
for row in ws1.values:
    for value in row:
        print(value, end='\t')
    print()

print('===='*10)


row1=ws1['1']    #시트 첫 행 복사, row1은 튜플이다.
for ele  in  row1:
	print(ele.value, end=', ')
print('\n---------')

important_date=[ '01/24/2013', '01/31/2013' ]

list_date=[]
list_date_str=[]
cost_pos=[]
for row in ws1.rows:
    for ele in row:
        if ele.data_type=='d':    # date, cell type
            COLUMN    =ele.column
            ROW           =ele.row
            list_date.append(ele.value)
            list_date_str.append( 
                ele.value.strftime(" %A, %d, %B, %Y ")    )
            print(ROW, COLUMN, ele.value, "$", ws1.cell(row=ROW, column=COLUMN-1).value)
            
            if ws1.cell(row=ROW,column=COLUMN-1).value > 1400:
                cost_pos.append( (ROW, COLUMN-1) )
        else:
            pass
            
print(cost_pos)
