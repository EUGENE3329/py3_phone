#!/usr/bin/python3

import openpyxl
from openpyxl import Workbook

import sys
import os

os.system('clear')
#excel_f=sys.argv[1]
excel_f ='sales_2013.xlsx'

if not os.path.exists(excel_f) or not os.path.isfile(excel_f):
    print('file open error')
    exit(1)

#load from file
wb=openpyxl.load_workbook(excel_f)
print('number of sheets in workbooks: ', len(wb.sheetnames) )
print(wb.sheetnames)    # value type is list
print('name\t\t', 'Rows', '\t','Columns')
for sheet in wb:        # extracting each sheet in wb 
    RO = 1
    COL = 1
    for ele in sheet.rows:
        RO+=1
    for ele in sheet.columns:
        COL+=1
    
    print( sheet.title, '\t' ,RO, '\t', COL)

print('===============================')