#!/usr/bin/python3

import os
from re import T
import sys

from datetime import date

import openpyxl

os.system('cls')

#작업 디렉토리가 실행파일 장소와 일치하지 않아서 생기는 에러 수정
working_dir=os.getcwd()
print('work here: ', working_dir )
if working_dir != "D:\개발_코딩연습\python_연습\데이터분석입문\ch3":
    os.chdir("D:\개발_코딩연습\python_연습\데이터분석입문\ch3")
print("Now !!!: ", os.getcwd())

#input_xl_f = sys.argv[1]
#input_xl_f = sys.argv[1]
input_xl_f =  'sales_2013.xlsx'
output_xl_f = 'tmp2.xlsx'

output_wb = openpyxl.Workbook()
output_wb.create_sheet('jan_2013_output')

#wb=openpyxl.load_workbook(input_xl_f, read_only=True, data_only=True)
wb=openpyxl.load_workbook(input_xl_f)
ws=wb['january_2013']

#경고 삭제, 아래 순환문 때문에 작성 함.
import warnings
warnings.simplefilter('ignore')

# 공식 문서 추천, 값만 추출
for row in ws.values:
    for value in row:
        print(value, end='\t')
    print()
    
#셀에 저장된 숫자, 값의 데이터 형식 출력
for row in ws.iter_rows():
    for column in ws.iter_cols():
        print( len(row), len(column) )
    print()
    print( row[:].value )





'''
# 작동 안 함. 연습 필요
for row_idx in ws.rows:
    row_list_output=[]
    for col_idx in ws.columns:
        print( ws.cell(row_idx, col_idx) )
    print()
'''

