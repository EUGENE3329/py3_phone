#!/usr/bin/python3
'''
특정 날짜가 저장된 행을 추출한다.
'''
import os
from re import T
import sys

from datetime import date, datetime, timedelta
#import datetime    #strptime(), 
                                      #네임스페이스  영역에 걸린다. 위 행과
                                      #논리적 충돌이 없게하라.

import openpyxl

os.system('clear')

#작업 디렉토리가 실행파일 장소와 일치하지 않아서 생기는 에러 수정
working_dir=os.getcwd()
print('work here: ', working_dir )

'''pc경로
if working_dir != "D:\개발_코딩연습\python_연습\데이터분석입문\ch3":
    os.chdir("D:\개발_코딩연습\python_연습\데이터분석입문\ch3")
print("Now !!!: ", os.getcwd())
'''

'''스마트폰 경로
/storage/emulated/0/Documents/py3/data_mgm_intro/p162_1.py
'''
if working_dir != "/storage/emulated/0/Documents/py3_phone/data_mgm_intro/":
    os.chdir("/storage/emulated/0/Documents/py3_phone/data_mgm_intro/")
print("Now !!!: ", os.getcwd())

input_xl_f =  'sales_2013.xlsx'
output_xl_f = 'tmp3.xlsx'

output_wb = openpyxl.Workbook()
#output_wb.create_sheet('jan_2013_output')
#output_ws1=output_wb.get_sheet_by_name('jan_2013_output')
output_ws1=output_wb.active
output_ws1.title='jan_2013_output'


#경고 삭제, 아래 순환문 때문에 작성 함.
#폰 에서 코딩 위치
import warnings
warnings.simplefilter('ignore')

wb=openpyxl.load_workbook(input_xl_f)
ws1=wb['january_2013']

for row in ws1.values:
    for value in row:
        print(value, end='\t')
    print()

print('===='*10)


row1=ws1['1']    #시트 첫 행 복사, row1은 튜플이다.
for ele  in  row1:
	print(ele.value, end=', ')
print('\n---------')


t_date1=[ '01/24/2013', '01/31/2013' ]
t_date2=[ '01/24/22', '01/31/22' ]
set_date=[t_date1, t_date2]

ws_date=[]
ws_date_str=[]
cost_pos=[]
print("시트에서 추출==========")
for row in ws1.rows:
    for ele in row:
        if ele.data_type=='d':    # date, cell type
            COLUMN    =ele.column
            ROW           =ele.row
            ws_date.append(ele.value)
            ws_date_str.append( 
                ele.value.strftime(" %a, %d, %B, %Y ")    )
                #(ex) Sun, 31, May, 2022
                #객체를 문자열로 출력
            print(ele.number_format)    #'mm-dd-yy'
            
        
print("검사###: strptime 연습")
for type in set_date:
    for ele in type:
            ea=ele.split('/')
            if len( ea[2] ) == 4:
                print( datetime.strptime(    \
                    ele, "%m/%d/%Y"    #for yyyy type, t_date1
                    )    )
            if len( ea[2] ) == 2:
                print( datetime.strptime(    \
                    ele, "%m/%d/%y"    #for yy type, t_date2
                    )    )

            
print('======='*5)
#print(cost_pos)
print("datetime객체->str: ", ws_date_str[0])
print("str->datetime객체: ", ws_date[0].strftime("%m,%d, %Y"))


