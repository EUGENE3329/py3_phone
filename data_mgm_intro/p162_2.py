#!/usr/bin/python3

import os
from re import T
import sys

from datetime import date

import openpyxl

os.system('clear')

#작업 디렉토리가 실행파일 장소와 일치하지 않아서 생기는 에러 수정
working_dir=os.getcwd()
print('work here: ', working_dir )

'''pc경로
if working_dir != "D:\개발_코딩연습\python_연습\데이터분석입문\ch3":
    os.chdir("D:\개발_코딩연습\python_연습\데이터분석입문\ch3")
print("Now !!!: ", os.getcwd())
'''

'''스마트폰 경로
/storage/emulated/0/Documents/py3/data_mgm_intro/p162_1.py
'''
if working_dir != "/storage/emulated/0/Documents/py3_phone/data_mgm_intro/":
    os.chdir("/storage/emulated/0/Documents/py3_phone/data_mgm_intro/")
print("Now !!!: ", os.getcwd())

#input_xl_f = sys.argv[1]
#input_xl_f = sys.argv[1]
input_xl_f =  'sales_2013.xlsx'
output_xl_f = 'tmp3.xlsx'

output_wb = openpyxl.Workbook()
#output_wb.create_sheet('jan_2013_output')
#output_ws1=output_wb.get_sheet_by_name('jan_2013_output')
output_ws1=output_wb.active
output_ws1.title='jan_2013_output'


#경고 삭제, 아래 순환문 때문에 작성 함.
#폰 에서 코딩 위치
import warnings
warnings.simplefilter('ignore')

#wb=openpyxl.load_workbook(input_xl_f, read_only=True, data_only=True)
wb=openpyxl.load_workbook(input_xl_f)
ws1=wb['january_2013']

'''
#pc에서 코딩 위치
import warnings
warnings.simplefilter('ignore')
'''

# 공식 문서 추천, 값만 추출
for row in ws1.values:
    for value in row:
        print(value, end='\t')
    print()

print('===='*10)

'''
#셀에 저장된 숫자, 값의 데이터 형식 출력
for row in ws1.iter_rows():
    print(row[:])
#    print( row[:].value )
'''

'''
셀 범위를 리턴하면 ( ws1['A3C3'] ), 튜플(2차원배열)이 리턴된다.
따라서, 이 튜플을 일반  변수로 저장한 후, 그 튜플의 원소들을 
제어하면, 시트 상의 셀들을 제어할 수 있다.
변경된 튜플, 즉 시트의 셀들을 저장하려면, 
해당 시트의 워크북을 저장하면 된다.
'''
row1=ws1['1']    #시트 첫 행 복사, row1은 튜플이다.
for ele  in  row1:
	print(ele.value, end=', ')
print('\n---------')

tmp_value=[
'고객 아뒤', 
'고객 이름', 
'주문 번호',
 '판매액', 
'구매일' 
]
# 리턴된 튜플의 값을 수정해서,  셀값을 수정할 수 있다.
for idx in range( len(tmp_value) ):
	row1[idx].value=tmp_value[idx]
	
for ele  in  row1:        #수정 됨 확인.
	print(ele.value, end=', ')

'''	
ans=input("D U want 2 save changing it? 'y' or 'n' ")
if  ans== 'y':
	wb.save(input_xl_f)
'''

from datetime import datetime

#os.system('clear')
print('extracting date & time ****')
list_date=[]
list_date_str=[]
for row in ws1.rows:
    for ele in row:
        if ele.data_type=='d':    # date cell type
            list_date.append(ele)
            list_date_str.append( 
                ele.value.strftime(" %A, %d, %B, %Y ")    )
        else:
            pass
            
'''
셀에 환율, 회계, 등의 데이터를 저장하고 위 순환문을 돌리면?
'''
print(list_date[0].value)
p=list_date[0].value

print(p.year, p.month)
print( p.strftime(" %A, %d, %B, %Y ") )
print(  "====="*10 )

for idx in range( len(list_date_str) ):
    print( list_date_str[idx] )
    output_ws1.cell(row=(idx+1),    \
                                    column=1,       \
                                    value = list_date_str[idx]
                                    )

ans=input("D U want 2 save changing it? 'y' or 'n' ")
if  ans== 'y':
    output_wb.save(output_xl_f)
