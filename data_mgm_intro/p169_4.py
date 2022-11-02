#!/usr/bin/python3

'''
목적
특정 날짜가 저장된 행을 추출한다.
    ver up) 문자열로 '추출할 날짜'를 입력받는다.

routine
1) 작동 시스템 확인, 작업폴더 확인 후 지정
    폰? 윈도우? 패드?
2) 입력 출력용 엑셀을 연다.
    sales_2013.xlsx    tmpXXX.xlsx
3)지정된 날짜로 추출한다.
    tmp에 저장한다.
'''

import platform, os

SYSTEM=platform.system()
MACHINE=platform.machine()
#작업폴더 설정
working_dir=os.getcwd()
if SYSTEM == "Windows":
    os.chdir("D:\개발_코딩연습\python_연습\데이터분석입문\ch3")
    pass
elif SYSTEM == "macOS":
    pass
elif SYSTEM== "Linux":
    if MACHINE == 'aarch64':    #smart phone
        os.chdir("/storage/emulated/0/Documents/py3_phone/data_mgm_intro/")
        
    pass
else:
    print('i have no idea what this system !!!')
    
print( 'working_dir : ', os.getcwd() )

import  openpyxl
#엑셀파일, 파일모드 필요없음.
xl_file_name = "sales_2013.xlsx"
dst_xl_name = "tmp3.xlsx"
if os.path.isfile( xl_file_name ):
    xl_wb= openpyxl.load_workbook( xl_file_name )
else:
	print( " xl file  open  error, not exist ! " )
                                                       
#ws = xl_wb[1]
print( 'sheet titles : ',   xl_wb.sheetnames )
sheets_dic= {}
paper=1
for wsTitle in   xl_wb.sheetnames:
    sheets_dic[ paper ] = wsTitle
    paper += 1
    
paperPick = input("sheet name ?, ( use number \
starting from '1' ) \n")
ws = xl_wb[  
        sheets_dic[ 
            int( paperPick)
            ]
        ]

'''
#checking
for ROW in  ws.rows:
    for CELL in ROW:
        print(   CELL.value , end='\t')
'''

# printing titles, header
print('    \n\n   ')
for cell in ws[ '1' ]:    
    print( cell.value, end='\t' )
print('\n')
# fetching dates, p166_3 참고
# finding the idx of column which type is date.
idx_colu_date = None
for cell in ws[ '2' ]:    #2행에서 날짜셀 찾기
    if cell.data_type == 'd':
        idx_colu_date = cell.column    #정수형
        print( idx_colu_date, cell.colu )

#for column in ws.columns:
#    print(  column.value  )

#ending routine
xl_wb.save( dst_xl_name )