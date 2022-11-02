#!/usr/bin/python3

import os

import openpyxl


os.system('clear')
xl_read_f='sales_2013.xlsx'
xl_write_f='tmp.xlsx'


output_wb=openpyxl.Workbook()
output_ws=output_wb.create_sheet('jan_2013_output')

with openpyxl.Workbook(  xl_read_f ) as input_wb:
    ws=input_wb['january_2013']
    for idx in ws.rows:
        print('ok')

