import os

from openpyxl import Workbook
import openpyxl

xl_name='tmp2_lotto.xlsx'
if os.path.exists(xl_name) and os.path.isfile(xl_name):
    wb=openpyxl.load_workbook(xl_name)
    print('file ok!!!')
    print(wb.sheetnames)
else:
    pass
    
'''
#removing sheets.
for ele in wb.sheetnames[2:]:
    wb.remove( wb[ele])

#wb.remove_sheet(wb.sheetnames[2])    #not working.
#wb.remove(wb['bbb1'])    # working.
print(wb.sheetnames)
wb.save(xl_name)
'''

ws1=wb[ wb.sheetnames[0] ]
ws1['A1']='999'
a=ws1['A1']
#not error, but not printing to console
print(a) 

'''
#cell test inputting value.
ws1.cell(row=1,column=1, value=' dfg ')
for  ele_column in range(1,7):
    #for_loop to make randum num.
    #    use a ran_fun() which return a list with values.
'''    
    
import random

numlist=range(1,46)    #making lottery nums
for ele_row in range(1,101):
    for  ele_column in range(1,7):
        s=random.sample(numlist,6)
        ws1.cell(row=ele_row,
                       column=ele_column, 
                       value=s[ ele_column-1 ]
                       )
                       
ans=input("do u want 2 save nums to file? , 'y' or 'yes'  ")
if ans=='y' or ans=='yes':
    wb.save(xl_name)
    
