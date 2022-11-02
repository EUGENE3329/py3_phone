import openpyxl


#input_f=sys.argv[1]
wb1= openpyxl.load_workbook("sup1.xlsx")

#will be  rejected method, "get_sheet_names()".
#tmp=wb1.get_sheet_names()    
tmp=wb1.sheetnames    #extracting all names in a workbook.
print(tmp,type(tmp))   #list type.


#will be  rejected method, "get_sheet_by_names()".
#ws4=wb1.get_sheet_by_name(tmp[0])
ws5=wb1[tmp[0]]    #extracting one sheet OBJ with a name.
#print(dir(ws5))     #print all methods.

#if ws4==ws5:
#    print('ttt')

# printing one sheet contents.
print('≈==============≈=')
print(
ws5.cell(row=1, column=1).value
)
for r in range(1,5):
    for c in range(1,5):
        print( ws5.cell(row=r, column=c).value, end=', ' )
    print()


print("==== making  a new xlsx")
wb=openpyxl.Workbook()
ws=wb.active    #getting a worksheet.

#making a sheet.
ws1=wb.create_sheet("my_sheet1")
ws2=wb.create_sheet("my_sheet2",0)   #첫버째 위치.
ws3=wb.create_sheet("my_sheet3", -1) #끝에서 드번째.

ws3.title="3333"

print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)
    
    
print("==≈=======≈")


#for line in ws4['1']:
#    print( line.value )

#for r in dataframe_to_rows( df, index=True, header=True):
    
import os.path
if os.path.isfile( "test_empty.xlsx" ):
    print("existing the file, !!!")
else:
    wb.save("test_empty.xlsx")