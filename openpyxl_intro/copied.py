from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb=Workbook()
dst_f='empty_book.xlsx'
ws1=wb.active
ws1.title='range names'

for row in range(1,40):
    ws1.append( range(600) )
    
    
ws2=wb.create_sheet(title='pi')
ws2['F5']=3.14

ws3=wb.create_sheet(title=' Data ')
for row in range(10,20):
    for col in range(27,54):
        ws3.cell(
             column=col,
             row=row,
             value="{0}".format( get_column_letter(col) )
        )

vv=1
for col in range(1,28):
                ws3.cell(column=col,    row=9,    value=vv        )
                vv+=1
        
print(ws3['AA10'].value)    # row=10,col=27(AA)


ans=input("do u want 2 save nums to file? , 'y' or 'yes'  ")
if ans=='y' or ans=='yes':
    wb.save(filename=dst_f)