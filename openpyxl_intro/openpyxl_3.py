'''ref)
https://openpyxl.readthedocs.io/en/stable/tutorial.html#create-a-workbook
'''

import openpyxl
from openpyxl import Workbook

'''
#making xlsx file.
wb=Workbook()
ws=wb.active

wb.save('tmp2.xlsx')
'''

import os

xl_name='tmp2.xlsx'
if os.path.exists(xl_name) and os.path.isfile( xl_name ):
    xl_fd=openpyxl.load_workbook(xl_name)
    print('!!! file open', end='\n----------\n')
else:
    print('no! the file.')

wb=xl_fd    #not using fd, use workbook.
print(wb.sheetnames)
ws1=wb.active
ws1.title='aaa'
print(ws1.title)
ws2=wb.create_sheet('bbb')
print(wb.sheetnames)
#wb.save(xl_name)

ws1['A1    ']='123'
ws1['A2']=123
wb.save(xl_name)

